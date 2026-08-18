"""
===============================================================
  MIS Excel Automation  —  Web App (Flask)
  Cloud-deployable | Login-protected | Mobile-friendly
  v3 — Enhanced Processing Engine
       • Smarter amount-column detection (name + data patterns)
       • Processes ALL detected amount columns automatically
       • In-place replacement: no stray duplicate columns
       • Mail-merge-safe plain-text formatted column (Indian commas)
       • Paise / decimal precision preserved; no scientific notation
       • Fully in-memory (works on Render free tier)
===============================================================
"""

import os, re, math, json, io
from datetime import datetime
from flask import (Flask, render_template, request, redirect,
                   url_for, session, send_file, flash, jsonify)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "mis-secret-key-change-in-prod-2024")

# In-memory stores (no disk needed)
FILE_STORE   = {}   # username -> {"data": bytes, "name": str}
OUTPUT_STORE = {}   # username -> {"data": bytes, "name": str}
ACTIVITY_LOG = []   # list of log dicts (in memory)

# ── Credentials ───────────────────────────────────────────────
USERS = {
    "achal13": {"password": "achal1331", "role": "admin"},
}

# ══════════════════════════════════════════════════════════════
# NUMBER HELPERS
# ══════════════════════════════════════════════════════════════

IFMT = '[>=10000000]##\\,##\\,##\\,##0;[>=100000]##\\,##\\,##0;##,##0'
# Variant that also shows up-to-2 decimal places (paise)
IFMT_DEC = '[>=10000000]##\\,##\\,##\\,##0.##;[>=100000]##\\,##\\,##0.##;##,##0.##'


def _safe_float(val):
    """Convert a cell value to float, returning None on failure."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    s = str(val).strip().replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def format_indian_number(n, with_paise=False):
    """
    Return Indian comma-formatted string e.g. 15,00,000 or 15,00,000.50.
    Uses plain text — safe for Word Mail Merge.
    """
    if n is None:
        return ""
    is_neg = n < 0
    n = abs(n)

    if with_paise and not n.is_integer():
        int_part = int(n)
        paise    = round((n - int_part) * 100)
        formatted_int = _indian_int_str(int_part)
        result = f"{formatted_int}.{paise:02d}"
    else:
        formatted_int = _indian_int_str(int(round(n)))
        result = formatted_int

    return ("-" if is_neg else "") + result


def _indian_int_str(n):
    """Convert non-negative integer to Indian-comma string."""
    if n == 0:
        return "0"
    s = str(n)
    if len(s) <= 3:
        return s
    result = s[-3:]
    s = s[:-3]
    while s:
        result = s[-2:] + "," + result
        s = s[:-2]
    return result


def num_to_words_indian(n):
    """Convert a number to Indian words, handling paise."""
    if n is None:
        return ""
    if isinstance(n, float) and math.isnan(n):
        return ""
    n = float(n)
    if n == 0:
        return "Zero Rupees Only"

    is_neg = n < 0
    n = abs(n)

    rupees = int(n)
    paise  = int(round((n - rupees) * 100))

    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight',
            'Nine', 'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen',
            'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
    tens_w = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty',
              'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def td(n):
        return ones[n] if n < 20 else tens_w[n // 10] + (' ' + ones[n % 10] if n % 10 else '')

    def thd(n):
        return (ones[n // 100] + ' Hundred' + (' ' + td(n % 100) if n % 100 else '')) \
            if n >= 100 else td(n)

    parts = []
    rem = rupees
    crore = rem // 10_000_000;  rem %= 10_000_000
    lakh  = rem // 100_000;     rem %= 100_000
    thou  = rem // 1_000;       rem %= 1_000
    if crore: parts.append(thd(crore) + ' Crore')
    if lakh:  parts.append(td(lakh)   + ' Lakh')
    if thou:  parts.append(td(thou)   + ' Thousand')
    if rem:   parts.append(thd(rem))

    rupee_words = ' '.join(parts) + ' Rupees' if parts else ''
    paise_words = td(paise) + ' Paise' if paise else ''

    if rupee_words and paise_words:
        result = rupee_words + ' and ' + paise_words + ' Only'
    elif rupee_words:
        result = rupee_words + ' Only'
    else:
        result = paise_words + ' Only'

    return ('Minus ' if is_neg else '') + result


# ══════════════════════════════════════════════════════════════
# ENHANCED AMOUNT-COLUMN DETECTION
# ══════════════════════════════════════════════════════════════

# Header keywords that strongly suggest an amount column
_AMT_KEYWORDS = [
    'amount', 'amt', 'total', 'arrear', 'dpi', 'principal', 'overdue',
    'disburse', 'outstanding', 'bounce', 'penalty', 'received', 'interest',
    'award', 'claim', 'sold', 'balance', 'loan', 'settlement', 'demand',
    'debit', 'credit', 'payment', 'repay', 'instalment', 'installment',
    'emi', 'dues', 'recovery', 'waiver', 'fee', 'charge', 'fine',
]

# Header patterns that should NOT be treated as amount columns
_SKIP_PATTERNS = [
    'in word', 'in words', 'formatted', 'date', 'name', 'no ',
    'number', 'address', 'status', 'remark', 'reason', 'url',
    'email', 'phone', 'mobile', 'code', 'id ', ' id', 'sr.',
    'serial', 'branch', 'district', 'state', 'pincode', 'zip',
    'gender', 'category', 'type', 'description', 'narration',
]

# Compiled regex: looks like a currency value (digits, optional commas/dots)
_CURRENCY_RE = re.compile(r'^\s*-?\s*[\d,]+(\.\d{1,2})?\s*$')


def _header_suggests_amount(header: str) -> bool:
    h = header.strip().lower()
    if any(s in h for s in _SKIP_PATTERNS):
        return False
    return any(k in h for k in _AMT_KEYWORDS)


def _column_has_numeric_data(ws, col_idx, sample_rows=10) -> bool:
    """Check if at least one cell in the first sample_rows data rows is numeric."""
    numeric_count = 0
    total_non_empty = 0
    for r in range(2, min(ws.max_row + 1, sample_rows + 2)):
        v = ws.cell(row=r, column=col_idx).value
        if v is None or str(v).strip() == '':
            continue
        total_non_empty += 1
        f = _safe_float(v)
        if f is not None:
            numeric_count += 1
    if total_non_empty == 0:
        return False
    return (numeric_count / total_non_empty) >= 0.7   # ≥70 % numeric


def find_amount_columns(ws):
    """
    Improved detection: returns list of (col_index, header_name).
    Uses both header-keyword matching AND data-pattern analysis.
    Also uses pure data-pattern detection for columns with no obvious header.
    """
    found = []
    seen_cols = set()

    # Pass 1 — header keyword match + numeric data check
    for cell in ws[1]:
        if not cell.value:
            continue
        h = str(cell.value).strip()
        if _header_suggests_amount(h) and _column_has_numeric_data(ws, cell.column):
            found.append((cell.column, h))
            seen_cols.add(cell.column)

    # Pass 2 — pure numeric columns with no keyword header
    #           (catches "Amt", "Rs.", "₹", unnamed money cols)
    for cell in ws[1]:
        if cell.column in seen_cols:
            continue
        if not cell.value:
            h = f"Column {get_column_letter(cell.column)}"
        else:
            h = str(cell.value).strip()
            hl = h.lower()
            # Skip clearly non-amount headers
            if any(s in hl for s in _SKIP_PATTERNS):
                continue
            # Skip headers that contain alpha text which clearly aren't money
            # but also skip already processed "in words" / "formatted" cols
            if any(k in hl for k in ['word', 'format']):
                continue

        # Stronger evidence needed when header doesn't match keywords
        if _column_has_numeric_data(ws, cell.column, sample_rows=15):
            # Additional check: values must look like currency amounts
            vals = []
            for r in range(2, min(ws.max_row + 1, 16)):
                v = ws.cell(row=r, column=cell.column).value
                if v is not None:
                    vals.append(v)
            currency_like = sum(
                1 for v in vals
                if isinstance(v, (int, float)) or _CURRENCY_RE.match(str(v))
            )
            if vals and (currency_like / len(vals)) >= 0.8:
                found.append((cell.column, h))
                seen_cols.add(cell.column)

    return found


def _has_decimal(ws, col_idx) -> bool:
    """Return True if any cell in the column has a non-zero fractional part."""
    for r in range(2, min(ws.max_row + 1, 50)):
        v = _safe_float(ws.cell(row=r, column=col_idx).value)
        if v is not None and not float(v).is_integer():
            return True
    return False


# ══════════════════════════════════════════════════════════════
# EXCEL PROCESSING ENGINE  (v3)
# ══════════════════════════════════════════════════════════════

def process_excel_bytes(input_bytes, rules):
    """
    Process Excel from bytes, return output as bytes.
    `rules` maps lowercase header -> [mode]  where mode ∈ {"words","fill","none"}

    Changes vs v2
    ─────────────
    • Amount columns are processed IN-PLACE (original column updated with Indian
      number format; no stale original-value column left behind).
    • A single plain-text "(Formatted)" helper column is inserted immediately to
      the right for Mail Merge — this contains an unambiguous text string.
    • If mode=="words", an "in Word" column is also inserted.
    • Existing "in Word" / "in Words" columns next to amount cols are filled
      (mode=="fill") rather than duplicated.
    • Large numbers never use scientific notation (stored as int/float with
      explicit number_format).
    • Paise are preserved; words include "X Rupees and Y Paise Only".
    """

    # ── Style presets ─────────────────────────────────────────
    af   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    afo  = Font(bold=True, color="276221", name="Calibri", size=10)
    wf   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wfo  = Font(bold=True, color="7F6000", name="Calibri", size=10)
    mf   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    mfo  = Font(bold=True, color="1F4E79", name="Calibri", size=10)
    hf   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfo  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ctr  = Alignment(horizontal='center',  vertical='center', wrap_text=True)
    rgt  = Alignment(horizontal='right',   vertical='center')
    wrap = Alignment(wrap_text=True,       vertical='center')

    wb = load_workbook(io.BytesIO(input_bytes), data_only=True)

    for sn in wb.sheetnames:
        ws = wb[sn]

        # ── Build match list (sorted RIGHT→LEFT so insertions don't shift indexes) ──
        matched = []
        for cell in ws[1]:
            if not cell.value:
                continue
            h  = str(cell.value).strip()
            hl = h.lower()
            if hl in rules:
                matched.append((cell.column, h, rules[hl][0]))

        if not matched:
            continue

        matched_sorted = sorted(matched, key=lambda x: x[0], reverse=True)

        for orig_ci, col_name, mode in matched_sorted:

            # Guard: if every value in this column is unreadable (e.g. a
            # formula with no cached result — file was never opened/saved
            # in real Excel), fail with a clear message instead of
            # silently skipping every row.
            numeric_rows = sum(
                1 for r in range(2, ws.max_row + 1)
                if _safe_float(ws.cell(row=r, column=orig_ci).value) is not None
            )
            if numeric_rows == 0:
                raise ValueError(
                    f"Column '{col_name}' has no readable numeric values. "
                    f"If this column contains formulas, please open the file "
                    f"in Excel, let it recalculate, save it, and re-upload."
                )

            # Detect whether this column ever has decimal values
            has_dec = _has_decimal(ws, orig_ci)
            num_fmt  = IFMT_DEC if has_dec else IFMT

            # ── 1. Format original column IN-PLACE ───────────────
            hdr_cell = ws.cell(row=1, column=orig_ci)
            hdr_cell.fill      = af
            hdr_cell.font      = afo
            hdr_cell.alignment = ctr
            ws.column_dimensions[hdr_cell.column_letter].width = 18

            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=orig_ci)
                v = _safe_float(c.value)
                if v is not None:
                    # Store as proper numeric type — avoids scientific notation
                    c.value        = v if has_dec else int(round(v))
                    c.number_format = num_fmt
                    c.alignment    = rgt

            # ── 2. Insert Mail-Merge-safe plain-text column ───────
            # Always insert immediately to the right of original column.
            # After insert, original is still at orig_ci; new col is orig_ci+1.
            ws.insert_cols(orig_ci + 1)
            mc      = orig_ci + 1
            mh_cell = ws.cell(row=1, column=mc)
            mh_cell.value      = col_name.strip() + " (Formatted)"
            mh_cell.fill       = mf
            mh_cell.font       = mfo
            mh_cell.alignment  = ctr
            ws.column_dimensions[mh_cell.column_letter].width = 22

            for r in range(2, ws.max_row + 1):
                src_v = _safe_float(ws.cell(row=r, column=orig_ci).value)
                if src_v is not None:
                    dest = ws.cell(row=r, column=mc)
                    dest.value     = format_indian_number(src_v, with_paise=has_dec)
                    dest.alignment = rgt
                    # Force text storage — critical for Mail Merge
                    dest.number_format = '@'

            # ── 3. Words column (mode="words" or mode="fill") ─────
            if mode == "words":
                # Insert a new "in Word" column to the right of Formatted col
                ws.insert_cols(mc + 1)
                wc      = mc + 1
                wh_cell = ws.cell(row=1, column=wc)
                wh_cell.value      = col_name.strip() + " in Word"
                wh_cell.fill       = wf
                wh_cell.font       = wfo
                wh_cell.alignment  = ctr
                ws.column_dimensions[wh_cell.column_letter].width = 55

                for r in range(2, ws.max_row + 1):
                    src_v = _safe_float(ws.cell(row=r, column=orig_ci).value)
                    if src_v is not None:
                        dest = ws.cell(row=r, column=wc)
                        dest.value         = num_to_words_indian(src_v)
                        dest.alignment     = wrap
                        dest.number_format = '@'

            elif mode == "fill":
                # Look for an existing "in word/words" column to the right
                wc = None
                for off in range(1, 5):
                    nc_val = ws.cell(row=1, column=mc + off).value
                    if nc_val and any(
                        k in str(nc_val).lower() for k in ['in word', 'in words']
                    ):
                        wc = mc + off
                        break
                if wc:
                    wh2 = ws.cell(row=1, column=wc)
                    wh2.fill      = wf
                    wh2.font      = wfo
                    wh2.alignment = ctr
                    ws.column_dimensions[wh2.column_letter].width = 55
                    for r in range(2, ws.max_row + 1):
                        src_v = _safe_float(ws.cell(row=r, column=orig_ci).value)
                        if src_v is not None:
                            dest = ws.cell(row=r, column=wc)
                            dest.value         = num_to_words_indian(src_v)
                            dest.alignment     = wrap
                            dest.number_format = '@'

        # ── 4. Style un-styled header cells ──────────────────────
        for cell in ws[1]:
            if cell.value:
                try:
                    rgb = cell.fill.fgColor.rgb \
                        if cell.fill.fgColor.type == 'rgb' else '00000000'
                except Exception:
                    rgb = '00000000'
                if rgb in ('00000000', 'FFFFFFFF', ''):
                    cell.fill      = hf
                    cell.font      = hfo
                    cell.alignment = ctr
        ws.row_dimensions[1].height = 40

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ══════════════════════════════════════════════════════════════
# ACTIVITY LOG
# ══════════════════════════════════════════════════════════════

def log_activity(user, action, detail=""):
    ACTIVITY_LOG.insert(0, {
        "time":   datetime.now().strftime("%d %b %Y  %H:%M:%S"),
        "user":   user,
        "action": action,
        "detail": detail
    })
    if len(ACTIVITY_LOG) > 200:
        ACTIVITY_LOG.pop()


# ══════════════════════════════════════════════════════════════
# AUTH HELPERS
# ══════════════════════════════════════════════════════════════

def current_user():
    return session.get("username")


def require_login(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def require_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        u = current_user()
        if not u or USERS.get(u, {}).get("role") != "admin":
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════
# ROUTES  (unchanged from v2 — all existing APIs intact)
# ══════════════════════════════════════════════════════════════

@app.route("/", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        user = USERS.get(u)
        if user and user["password"] == p:
            session["username"] = u
            session["role"]     = user["role"]
            log_activity(u, "Login", "Successful login")
            return redirect(url_for("dashboard"))
        else:
            error = "Invalid username or password."
            log_activity(u or "unknown", "Login Failed", "Bad credentials")
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    u = current_user()
    if u:
        log_activity(u, "Logout", "")
        FILE_STORE.pop(u, None)
        OUTPUT_STORE.pop(u, None)
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@require_login
def dashboard():
    return render_template("dashboard.html",
                           username=current_user(),
                           role=session.get("role", "user"))


@app.route("/detect-columns", methods=["POST"])
@require_login
def detect_columns():
    f = request.files.get("file")
    if not f or not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload a valid .xlsx or .xls file."}), 400

    file_bytes = f.read()
    fname      = f.filename
    FILE_STORE[current_user()] = {"data": file_bytes, "name": fname}

    try:
        wb   = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws   = wb.active
        cols = find_amount_columns(ws)
        wb.close()
        return jsonify({"columns": [{"col": c, "name": n} for c, n in cols]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/process", methods=["POST"])
@require_login
def process():
    data  = request.json
    rules = data.get("rules", {})
    u     = current_user()
    store = FILE_STORE.get(u)

    if not store:
        return jsonify({"error": "No file found. Please upload your file again."}), 400
    if not rules:
        return jsonify({"error": "No columns selected."}), 400

    fname     = store["name"]
    name, ext = os.path.splitext(fname)
    out_name  = name + "_UPDATED" + ext

    try:
        out_bytes = process_excel_bytes(store["data"], rules)
        OUTPUT_STORE[u] = {"data": out_bytes, "name": out_name}
        log_activity(u, "Processed File", fname)
        return jsonify({"success": True, "filename": out_name})
    except Exception as e:
        log_activity(u, "Process Error", str(e))
        return jsonify({"error": str(e)}), 500


@app.route("/download")
@require_login
def download():
    u     = current_user()
    store = OUTPUT_STORE.get(u)
    if not store:
        flash("No processed file found. Please process a file first.")
        return redirect(url_for("dashboard"))
    log_activity(u, "Downloaded File", store["name"])
    return send_file(
        io.BytesIO(store["data"]),
        as_attachment=True,
        download_name=store["name"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin")
@require_login
@require_admin
def admin():
    users_info = [{"username": u, "role": d["role"]} for u, d in USERS.items()]
    return render_template("admin.html",
                           username=current_user(),
                           logs=ACTIVITY_LOG,
                           users=users_info)


@app.route("/admin/add-user", methods=["POST"])
@require_login
@require_admin
def add_user():
    u = request.form.get("username", "").strip()
    p = request.form.get("password", "").strip()
    r = request.form.get("role", "user")
    if u and p and u not in USERS:
        USERS[u] = {"password": p, "role": r}
        log_activity(current_user(), "Added User", u)
        flash(f"User '{u}' added successfully.")
    else:
        flash("Username already exists or invalid input.")
    return redirect(url_for("admin"))


@app.route("/admin/remove-user/<username>")
@require_login
@require_admin
def remove_user(username):
    if username == current_user():
        flash("You cannot remove yourself.")
    elif username in USERS:
        del USERS[username]
        log_activity(current_user(), "Removed User", username)
        flash(f"User '{username}' removed.")
    return redirect(url_for("admin"))


@app.route("/admin/clear-logs")
@require_login
@require_admin
def clear_logs():
    ACTIVITY_LOG.clear()
    log_activity(current_user(), "Cleared Logs", "")
    return redirect(url_for("admin"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
